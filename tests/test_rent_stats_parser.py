from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.rent_stats.parser import (
    RentStatsFormatError,
    parse_nsw_lodgements,
    parse_vic_moving_annual,
)

FIXTURES = Path(__file__).parent / "fixtures" / "rent_stats"
NSW = (FIXTURES / "nsw_lodgements_sample.xlsx").read_bytes()
VIC = (FIXTURES / "vic_moving_annual_sample.xlsx").read_bytes()


def test_nsw_parses_clean_rows_and_normalises_dwellings():
    parsed = parse_nsw_lodgements(NSW)
    assert len(parsed.rows) == 20
    first = parsed.rows[0]
    assert (first.period, first.postcode, first.dwelling_type) == ("2026-07", "2000", "unit")
    assert first.bedrooms == 0 and first.weekly_rent == Decimal(290)
    assert {r.dwelling_type for r in parsed.rows} == {"unit", "house", "townhouse", "other"}


def test_nsw_counts_dirty_rows():
    parsed = parse_nsw_lodgements(NSW)
    assert parsed.skipped_rows == 2
    assert parsed.unknown_dwelling == 1
    unknown = [r for r in parsed.rows if r.postcode == "2000" and r.dwelling_type == "other"]
    assert [r.weekly_rent for r in unknown] == [Decimal(800)]


def test_nsw_header_guard_trips():
    wb = openpyxl.load_workbook(FIXTURES / "nsw_lodgements_sample.xlsx")
    wb.worksheets[0].cell(row=3, column=5, value="Rent")
    mutated = _bytes(wb)
    with pytest.raises(RentStatsFormatError, match="header"):
        parse_nsw_lodgements(mutated)


def test_vic_parses_periods_sheets_and_suppressions():
    stats = parse_vic_moving_annual(VIC)
    two_bed = [s for s in stats if s.dwelling_type == "unit" and s.bedrooms == 2]
    albert = [s for s in two_bed if s.area_code == "Albert Park-Middle Park-West St Kilda"]
    periods = sorted(s.period for s in albert)
    assert periods[0] == "2024-Q2" and periods[-1] == "2025-Q3" and len(periods) == 6
    latest = next(s for s in albert if s.period == "2025-Q3")
    assert (latest.median, latest.sample_size) == (Decimal(643), 144)
    assert all(s.area_code != "Group Total" for s in stats)
    all_props = [s for s in stats if s.dwelling_type == "all"]
    assert all_props and all(s.bedrooms is None for s in all_props)


def test_vic_missing_sheet_trips_guard():
    wb = openpyxl.load_workbook(FIXTURES / "vic_moving_annual_sample.xlsx")
    wb.remove(wb["3 bedroom house"])
    with pytest.raises(RentStatsFormatError, match="sheet"):
        parse_vic_moving_annual(_bytes(wb))


def _bytes(wb) -> bytes:
    import io

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_nsw_trailing_empty_rows_are_skipped():
    """openpyxl yields () for trailing blank rows in some real annual files (2023)."""
    from datetime import date

    from app.rent_stats import parser

    rows = iter(
        [
            (None,) * 5,
            (None,) * 5,
            parser.NSW_HEADER,
            (date(2023, 3, 1), 2000, "F", "1", "600"),
            (),
            (None, None),
        ]
    )
    parsed = parser._parse_nsw_rows(rows)
    assert len(parsed.rows) == 1 and parsed.skipped_rows == 0
