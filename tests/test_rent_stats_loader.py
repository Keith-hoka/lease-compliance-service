from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from sqlalchemy import func, select

from app.models import RentBondLodgement, RentStatistic
from app.rent_stats.loader import load_nsw_file, load_vic_file

FIXTURES = Path(__file__).parent / "fixtures" / "rent_stats"
NSW = (FIXTURES / "nsw_lodgements_sample.xlsx").read_bytes()
VIC = (FIXTURES / "vic_moving_annual_sample.xlsx").read_bytes()
URL = "https://example.test/src.xlsx"


async def _count(session, model):
    return (await session.execute(select(func.count()).select_from(model))).scalar_one()


async def _stat(session, **where):
    stmt = select(RentStatistic).filter_by(**where)
    return (await session.execute(stmt)).scalar_one()


async def _stat_count(session, **where):
    stmt = select(func.count()).select_from(RentStatistic).filter_by(**where)
    return (await session.execute(stmt)).scalar_one()


async def test_nsw_load_inserts_detail_and_aggregates(db_session):
    result = await load_nsw_file(db_session, "nsw_july_2026.xlsx", NSW, URL)
    await db_session.commit()
    assert (result.loaded_rows, result.skipped_rows, result.unknown_dwelling) == (20, 2, 1)
    assert result.periods == ["2026-07"] and result.unchanged is False
    assert await _count(db_session, RentBondLodgement) == 20

    unit_2000 = await _stat(
        db_session,
        jurisdiction="NSW",
        period="2026-07",
        area_code="2000",
        dwelling_type="unit",
        bedrooms=None,
    )
    assert unit_2000.sample_size == 6
    assert unit_2000.median == Decimal(760)
    assert unit_2000.p25 == Decimal("697.5") and unit_2000.p75 == Decimal("886.25")

    house_2150 = await _stat(
        db_session,
        jurisdiction="NSW",
        period="2026-07",
        area_code="2150",
        dwelling_type="house",
        bedrooms=None,
    )
    assert house_2150.sample_size == 5 and house_2150.median == Decimal(660)

    all_2000 = await _stat(
        db_session,
        jurisdiction="NSW",
        period="2026-07",
        area_code="2000",
        dwelling_type="all",
        bedrooms=None,
    )
    assert all_2000.sample_size == 9


async def test_nsw_reload_same_hash_is_noop(db_session):
    await load_nsw_file(db_session, "nsw_july_2026.xlsx", NSW, URL)
    await db_session.commit()
    again = await load_nsw_file(db_session, "nsw_july_2026.xlsx", NSW, URL)
    await db_session.commit()
    assert again.unchanged is True
    assert await _count(db_session, RentBondLodgement) == 20


async def test_nsw_changed_hash_replaces_file_rows(db_session):
    await load_nsw_file(db_session, "nsw_july_2026.xlsx", NSW, URL)
    await db_session.commit()
    trimmed = _drop_last_row(NSW)
    result = await load_nsw_file(db_session, "nsw_july_2026.xlsx", trimmed, URL)
    await db_session.commit()
    assert result.unchanged is False and result.loaded_rows == 19
    assert await _count(db_session, RentBondLodgement) == 19


async def test_nsw_reload_with_different_period_drops_stale_statistics(db_session):
    """A corrected re-upload that moves rows to a new period must not leave the old period's
    statistics behind - aggregate_nsw recomputes both the old and the new period."""
    await load_nsw_file(db_session, "nsw_july_2026.xlsx", NSW, URL)
    await db_session.commit()
    shifted = _shift_month(NSW, 6)
    result = await load_nsw_file(db_session, "nsw_july_2026.xlsx", shifted, URL)
    await db_session.commit()
    assert result.unchanged is False
    assert await _stat_count(db_session, jurisdiction="NSW", period="2026-07") == 0
    assert await _stat_count(db_session, jurisdiction="NSW", period="2026-06") > 0


FILE_B_ROWS = [
    (date(2026, 7, 5), 2000, "F", "1", "500"),
    (date(2026, 7, 6), 2000, "F", "1", "600"),
    (date(2026, 7, 7), 2000, "F", "1", "700"),
]


async def test_nsw_shared_period_recomputes_across_files(db_session):
    """Two files covering the same period contribute to one statistic (no source_file filter
    in aggregate_nsw's SQL); reloading one of them unchanged must not disturb the other's rows."""
    await load_nsw_file(db_session, "file_a.xlsx", NSW, URL)
    await db_session.commit()
    result_b = await load_nsw_file(db_session, "file_b.xlsx", _only_rows(FILE_B_ROWS), URL)
    await db_session.commit()
    assert result_b.loaded_rows == 3
    stat = await _stat(
        db_session,
        jurisdiction="NSW",
        period="2026-07",
        area_code="2000",
        dwelling_type="unit",
        bedrooms=None,
    )
    assert stat.sample_size == 9
    assert stat.median == Decimal(700)

    again = await load_nsw_file(db_session, "file_a.xlsx", NSW, URL)
    await db_session.commit()
    assert again.unchanged is True
    stat = await _stat(
        db_session,
        jurisdiction="NSW",
        period="2026-07",
        area_code="2000",
        dwelling_type="unit",
        bedrooms=None,
    )
    assert stat.sample_size == 9


async def test_vic_load_upserts_published_medians(db_session):
    result = await load_vic_file(db_session, "vic_sep_2025.xlsx", VIC, URL)
    await db_session.commit()
    assert result.unchanged is False and result.loaded_rows > 0
    stat = await _stat(
        db_session,
        jurisdiction="VIC",
        period="2025-Q3",
        area_code="Albert Park-Middle Park-West St Kilda",
        dwelling_type="unit",
        bedrooms=2,
    )
    assert (stat.median, stat.sample_size, stat.p25, stat.p75) == (Decimal(643), 144, None, None)
    again = await load_vic_file(db_session, "vic_sep_2025.xlsx", VIC, URL)
    assert again.unchanged is True


def _drop_last_row(data: bytes) -> bytes:
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.worksheets[0]
    ws.delete_rows(ws.max_row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _shift_month(data: bytes, month: int) -> bytes:
    """A copy of the workbook with every data row's lodgement date moved to `month`."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.worksheets[0]
    for (cell,) in ws.iter_rows(min_row=4, max_col=1):
        if isinstance(cell.value, datetime):
            cell.value = cell.value.replace(month=month)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _only_rows(rows: list[tuple]) -> bytes:
    """A copy of the NSW fixture's title/header rows with its data rows replaced by `rows`."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(NSW))
    ws = wb.worksheets[0]
    ws.delete_rows(4, ws.max_row - 3)
    for offset, values in enumerate(rows):
        for col, value in enumerate(values, start=1):
            ws.cell(row=4 + offset, column=col, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
