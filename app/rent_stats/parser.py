"""Workbook parsers for the official rent datasets, with fail-loud format guards."""

import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

import openpyxl

NSW_HEADER = ("Lodgement Date", "Postcode", "Dwelling Type", "Bedrooms", "Weekly Rent")
NSW_DWELLING = {"F": "unit", "H": "house", "T": "townhouse", "O": "other", "U": "other"}
VIC_SHEETS = {
    "1 bedroom flat": ("unit", 1),
    "2 bedroom flat": ("unit", 2),
    "3 bedroom flat": ("unit", 3),
    "2 bedroom house": ("house", 2),
    "3 bedroom house": ("house", 3),
    "4 bedroom house": ("house", 4),
    "All properties": ("all", None),
}
_VIC_PERIOD = re.compile(r"^(Mar|Jun|Sep|Dec) (\d{4})$")
_QUARTER = {"Mar": "Q1", "Jun": "Q2", "Sep": "Q3", "Dec": "Q4"}


class RentStatsFormatError(ValueError):
    """The source workbook does not match the pinned layout."""


@dataclass(frozen=True)
class Lodgement:
    period: str
    postcode: str
    dwelling_type: str
    bedrooms: int
    weekly_rent: Decimal


@dataclass(frozen=True)
class NswParse:
    rows: list[Lodgement]
    skipped_rows: int
    unknown_dwelling: int


@dataclass(frozen=True)
class VicStat:
    period: str
    area_code: str
    dwelling_type: str
    bedrooms: int | None
    median: Decimal
    sample_size: int


def parse_nsw_lodgements(data: bytes) -> NswParse:
    sheet = openpyxl.load_workbook(io.BytesIO(data), read_only=True).worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    next(rows), next(rows)
    header = tuple(next(rows)[:5])
    if header != NSW_HEADER:
        raise RentStatsFormatError(f"NSW header mismatch: {header}")
    parsed: list[Lodgement] = []
    skipped = unknown = 0
    for row in rows:
        lodged, postcode, dwelling, bedrooms, rent = row[:5]
        if lodged is None:
            continue
        try:
            beds = int(str(bedrooms))
            weekly = Decimal(str(rent))
        except (ValueError, InvalidOperation):
            skipped += 1
            continue
        code = str(dwelling)
        if code not in NSW_DWELLING:
            unknown += 1
        parsed.append(
            Lodgement(
                period=lodged.strftime("%Y-%m"),
                postcode=str(postcode),
                dwelling_type=NSW_DWELLING.get(code, "other"),
                bedrooms=beds,
                weekly_rent=weekly,
            )
        )
    return NswParse(rows=parsed, skipped_rows=skipped, unknown_dwelling=unknown)


def parse_vic_moving_annual(data: bytes) -> list[VicStat]:
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    missing = [name for name in VIC_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise RentStatsFormatError(f"VIC sheet missing: {missing}")
    stats: list[VicStat] = []
    for name, (dwelling, bedrooms) in VIC_SHEETS.items():
        stats.extend(_parse_vic_sheet(workbook[name], dwelling, bedrooms))
    return stats


def _parse_vic_sheet(sheet, dwelling: str, bedrooms: int | None) -> list[VicStat]:
    rows = list(sheet.iter_rows(values_only=True))
    periods = _vic_periods(rows[1], rows[2])
    out: list[VicStat] = []
    for row in rows[3:]:
        area = row[1]
        if not area or area == "Group Total":
            continue
        for period, count_col in periods:
            count, median = row[count_col], row[count_col + 1]
            if count in (None, "-") or median in (None, "-"):
                continue
            out.append(
                VicStat(
                    period=period,
                    area_code=str(area),
                    dwelling_type=dwelling,
                    bedrooms=bedrooms,
                    median=Decimal(str(median)),
                    sample_size=int(count),
                )
            )
    return out


def _vic_periods(label_row, kind_row) -> list[tuple[str, int]]:
    """(period, count-column index) pairs; guards the Count/Median alternation."""
    periods: list[tuple[str, int]] = []
    for col in range(2, len(label_row), 2):
        label = label_row[col]
        if label is None:
            break
        match = _VIC_PERIOD.match(str(label).strip())
        if not match or (kind_row[col], kind_row[col + 1]) != ("Count", "Median"):
            raise RentStatsFormatError(f"VIC header mismatch at column {col}: {label}")
        periods.append((f"{match.group(2)}-{_QUARTER[match.group(1)]}", col))
    if not periods:
        raise RentStatsFormatError("VIC header carries no periods")
    return periods
